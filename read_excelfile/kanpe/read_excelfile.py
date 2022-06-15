# PythonでExcelファイルを読み込んで、CSVファイルに書き出してみる
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

HANDSON_DIR = Path(__file__).parent
EXAMPLE_EXCEL_FILEPATH = HANDSON_DIR / "./example_tyohyo.xlsx"


def main() -> None:
    # Excelの数式はすでに計算された結果が入っていることを想定
    ex_wb = load_workbook(filename=EXAMPLE_EXCEL_FILEPATH, data_only=True)

    print(f"ワークシート一覧: {ex_wb.sheetnames}")

    # シート一つなので activeで有効にします。
    ex_ws = ex_wb.active
    ex_ws_value = ex_ws["A1"].value

    print(f"セルA1の値: {ex_ws_value}")

    # 読み込みたいセル番号を指定してみましょう
    celladdr = {
        "employee_number": "B3",
        "employee_name": "B4",
        "statement_number": "E3",
        "application_day": "E4",
        "total_amount": "D6",
    }

    cell_employee_number = ex_ws[celladdr["employee_number"]]
    cell_employee_name = ex_ws[celladdr["employee_name"]]
    cell_statement_number = ex_ws[celladdr["statement_number"]]
    # 注意: Excelで日付として表現されている値はdatetimeオブジェクトに変換されます
    cell_application_day = ex_ws[celladdr["application_day"]]
    cell_total_amount = ex_ws[celladdr["total_amount"]]

    print(
        f"""
    従業員番号:{cell_employee_number.value}
    申請者:{cell_employee_name.value}
    明細書番号:{cell_statement_number.value}
    申請日:{cell_application_day.value.isoformat(" ")}
    合計金額:{cell_total_amount.value}
    """
    )

    # セル範囲を指定して読み込んでみましょう

    cell_koumoku_list = ex_ws["A9":"D23"]
    # from pprint import pprint
    # pprint(cell_koumoku_list)
    column_names = [
        "日付",
        "内容",
        "支払先",
        "金額",
        "備考",
    ]

    for row in cell_koumoku_list:
        for column_name, column_cell in zip(column_names, row):
            if column_cell.value is not None:
                print(f"{column_name}:{column_cell.value},")


if __name__ == "__main__":
    main()
