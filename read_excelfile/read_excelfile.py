# PythonでExcelファイルを読み込んで、CSVファイルに書き出してみる
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

HANDSON_DIR = Path(__file__).parent
EXAMPLE_EXCEL_FILEPATH = HANDSON_DIR / "./example_tyohyo.xlsx"


def main() -> None:
    # Excelの数式はすでに計算された結果が入っていることを想定
    ex_wb = load_workbook(filename=EXAMPLE_EXCEL_FILEPATH, data_only=True)

    print(f"ワークシート一覧: {ex_wb.sheetnames}")

    # シート一つなので activeで有効にします。
    ex_ws = ex_wb.active
    ex_ws_value = ex_ws["A1"].value

    print(f"セルA1の値: {ex_ws_value}")
    # 続きの実装はハンズオン中に進めます


if __name__ == "__main__":
    main()
